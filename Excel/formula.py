from openpyxl import load_workbook
from openpyxl import Workbook
import datetime

wb = Workbook()
ws = wb.active

ws["A1"] = datetime.datetime.today()
ws["A2"] = "=SUM(1,2,3)" #1 + 2 + 3 설정
ws["A3"] = "=AVERAGE(1,2,3)"
ws["A4"] = "=SUM(A2:A3)"

wb.save("sample_formula.xlsx")