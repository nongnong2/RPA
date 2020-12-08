from openpyxl import load_workbook
# wb = load_workbook("sample_formula.xlsx")
# ws = wb.active

# # 수식 그대로 가져옴
# for row in ws.values:
#     for cell in row:
#         print(cell)

# 수식이 계산된 데이터로 가져오기 
wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active

# 수식 그대로 가져옴
#evaluate -> 아직 계산되지 않은건 None으로 뜸, 파일 생성이후 열어서 저장하고 실행 
for row in ws.values:
    for cell in row:
        print(cell)