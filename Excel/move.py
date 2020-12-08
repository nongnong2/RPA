from openpyxl import load_workbook

wb = load_workbook("sample4.xlsx")
ws = wb.active

#번호 영어 수학 -> 번호 (국어) 영어 수학 

ws.move_range("B1:C11", rows=0, cols=1) #B1:C11을 cols 1칸 이동 
ws["B1"].value = "국어"

wb.save("sample_move.xlsx")