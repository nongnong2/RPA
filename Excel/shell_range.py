from re import L
from openpyxl import Workbook
from random import *
from openpyxl.utils.cell import coordinate_from_string, coordinate_to_tuple

wb = Workbook()
ws = wb.active

ws.append(["번호", "영어", "수학"]) #1줄씩 데이터 넣기
for i in range(1, 11):
    ws.append([i,randint(0,100), randint(0,100)])

# col_B = ws["B"] #영어 column만 가져오기
# for cell in col_B:
#     print(cell.value)

# col_range = ws["B:C"] #영어, 수학 column 가져오기 
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

# row_title = ws[1]
# for cell in row_title:
#     print(cell.value)

# row_range = ws[1:6] # 1번쨰 줄에서 6번째 줄 가져오기
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()

# row_range = ws[2:ws.max_row] #2번쨰 줄부터 마지막 줄까지
# for rows in row_range:
#     for cell in rows:
#         # print(cell.value, end=" ")
#         # print(cell.coordinate, end=" ")
#         xy = coordinate_from_string(cell.coordinate)
#         # print(xy, end=" ")
#         print(xy[0], end=" ")
#         print(xy[1], end=" ")
#     print()

# #전체 rows
# print(tuple(ws.rows))

# for rows in tuple(ws.rows):
#     print(rows[0].value)

# #전체 columns
# print(tuple(ws.columns))
# for columns in tuple(ws.columns):
#     print(columns[0].value)

# #1번째 줄부터 11번쨰 줄까지, 2번째 열부터 3번쨰 열까지 
# for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3):
#     print(row[0].value, row[1].value) # 수학, 영어

for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):
    print(col[0].value, col[1].value)

wb.save("sample4.xlsx")