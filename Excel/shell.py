from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active
ws.title = "NadoSheet"

ws["A1"] = 1
ws["A2"] = 2

print(ws["A1"]) # A1 셀의 정보를 출력 
print(ws["A1"].value) # A1 셀의 값을 출력(값이 없으면 None 출력 )

print(ws.cell(row=1, column=1).value)

c = ws.cell(column=10, row=1, value=10) #ws["C1"] = 10
print(c.value)

#반복문을 이용해 랜덤 숫자 채우기
for x in range(1, 11): #10개 row
    for y in range(1, 11): #10개 column
        ws.cell(row=x, column=y, value=randint(0,100))

wb.save("sample3.xlsx")