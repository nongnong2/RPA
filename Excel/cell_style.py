from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fills import PatternFill
wb = load_workbook("sample4.xlsx")
ws = wb.active

#번호, 영어, 수학
a1 = ws["A1"]
b1 = ws["b1"]
c1 = ws["c1"]

ws.column_dimensions["A"].width = 5 #A열의 너비를 5로 설정
ws.row_dimensions[1].height = 50 #1번 줄의 높이를 50으로 지정 

# 스타일 적용
a1.font = Font(color="FF0000", italic=True, bold=True) # 글자 색은 빨강, Italic, 두껍게 
b1.font = Font(color="CC33FF", name="Arial", strike=True) #폰트 Arial, 글자 중간에 선 긋기 
c1.font = Font(color="0000FF", size=20, underline="single")# 글자크기 20, 밑줄 적용

# 테두리 적용
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 90점 넘는 셀에 대해 초록색 적용
for row in ws.rows:
    for cell in row:
        # 각 cell에 대해 정렬
        cell.alignment = Alignment(horizontal="center", vertical="center")

        if cell.column == 1: #A열 번호 제외
            continue
        # cell이 정수형이고 90점 보다 크면
        if isinstance(cell.value, int) and cell.value > 90:
            cell.fill = PatternFill(fgColor="00FF00", fill_type="solid")
            cell.font = Font(color="FF0000")

#틀 고정
ws.freeze_panes = "B2" #B2 기준으로 틀 고정 

wb.save("sample_cell_style.xlsx")