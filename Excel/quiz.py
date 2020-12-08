# 출석 10
# 퀴즈 1 - 10
# 퀴즈2 - 10
# 중간고사 - 20
# 기말고사 - 30
# 프로젝트 - 20

# 조건
# 1. 퀴즈2 점수를 10으로 수정
# 2. H열에 총점(SUM 이용), I열에 성적 정보 추가 
#  - 총점 90 이상 A, 80 이상 B, 70 이상 C, 나머지 D
#  3. 출석이 5미만인 학생은 총점 상관없이 F

# [현재까지 작성된 최종 성적 데이터 ]
# 학번, 출석, 퀴즈1, 퀴즈2, 중간고사, 기말고사, 프로젝트
# 1,10,8,5,14,26,12
# 2,7,3,7,15,24,18
# 3,9,5,8,8,12,4
# 4,7,8,7,17,21,18
# 5,7,8,7,16,25,15
# 6,3,5,8,8,17,0
# 7,4,9,10,16,27,18
# 8,6,6,6,15,19,17
# 9,10,10,9,19,30,19
# 10,9,8,8,20,25,20

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.alignment import Alignment

wb = Workbook()
ws = wb.active

thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

# 점수 데이터 집어넣기 
#max_col = 7, max_row =  10
max_col = 7
max_row = 10
values = [1,10,8,5,14,26,12,2,7,3,7,15,24,18,3,9,5,8,8,12,4,4,7,8,7,17,21,18,5,7,8,7,16,25,15,6,3,5,8,8,17,0,7,4,9,10,16,27,18,8,6,6,6,15,19,17,9,10,10,9,19,30,19,10,9,8,8,20,25,20]
i = 0 
for x in range(1, max_row+1):
    for y in range(1, max_col+1):
        cells = ws.cell(row=x, column=y, value=values[i])
        i += 1


# rows=1에 학번, 출석, 퀴즈1, 퀴즈2, 중간고사, 기말고사, 프로젝트 삽입하기 
ws.insert_rows(1)
columns = ["학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"]
for y in range(1, 8): 
    ws.cell(row=1,column=y, value=columns[y-1])

#퀴즈 2(column = 4)의 점수를 10으로 다 변경 
for row in ws.iter_rows(min_row=2, max_row=11, min_col=4, max_col=4):
    row[0].value = 10

# H열에 총점(SUM 이용), I열에 성적 정보 추가 
#  - 총점 90 이상 A, 80 이상 B, 70 이상 C, 나머지 D, 출석이 5미만인 학생은 총점 상관없이 F

ws["H1"] = "총점"
i = 2
for row in ws.iter_rows(min_row=2, max_row=11, min_col=8, max_col=8):
    row[0].value ="=SUM(B{0}:G{1})".format(i,i) 
    i += 1

ws["I1"] = "학점"
for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=9):
    
    total_sum = 0
    for i in range(0,6):
        total_sum += int(row[i].value)
    
    attendance_score =int(row[0].value) 

    if total_sum >= 90:
        row[7].value = "A"
    elif total_sum >= 80:
        row[7].value = "B"
    elif total_sum >= 70:
        row[7].value = "C"
    else: 
        row[7].value = "D"

    if attendance_score < 5:
        row[7].value = "F"

#셀 스타일 설정 
for row in ws.rows:
    for cell in row:
        cell.font = Font(color="FF0000", italic=True, bold=True) # 셀 전체 폰트 설정 
        cell.border = thin_border #테두리 설정 
        cell.alignment = Alignment(horizontal="center", vertical="center") #가운데 정렬 

wb.save("quiz.xlsx")
