from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() #새로운 Sheet 기본이름으로 생성
ws.title = "MySheet" #Sheet 이름 변경
ws.sheet_properties.tabColor = "ff66ff" #RGB형식으로 넣으면 sheet 색깔 변경 

ws1 = wb.create_sheet("YourSheet")
ws2 = wb.create_sheet("NewSheet",2) #2번쨰 index에 NewSheet 생성

new_ws = wb["NewSheet"] # Dict형태로 Sheet에 접근

print(wb.sheetnames) #모든 Sheet 이름 확인 

# Sheet 복사
new_ws["A1"] = "Test" #A1쉘에 Test 입력 
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample2.xlsx")

